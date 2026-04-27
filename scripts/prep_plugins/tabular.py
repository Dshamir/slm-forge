"""
Tabular plugin — CSV, TSV, XLSX, ODS, Parquet.

Each non-empty row becomes one chunk. If the table has a header, columns
are interleaved into "col: value" form to preserve semantic labels.

For Excel files with multiple sheets: each sheet contributes its own
chunks with chunk_idx counting per-sheet.
"""
from __future__ import annotations

import sys
from pathlib import Path
from typing import Iterator

from .orchestration_helpers import clean_text


def _row_to_text(headers, row) -> str:
    parts = []
    if headers and len(headers) == len(row):
        for h, v in zip(headers, row):
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            parts.append(f"{h}: {v}")
    else:
        parts = [str(v) for v in row if v is not None and str(v).strip()]
    return " — ".join(parts)


def _iter_csv(path: Path, delimiter: str, source_format: str):
    import csv
    try:
        with open(path, newline="", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f, delimiter=delimiter)
            try:
                headers = next(reader)
            except StopIteration:
                return
            for i, row in enumerate(reader):
                text = _row_to_text(headers, row)
                if text.strip():
                    yield i, text
    except Exception as e:
        sys.stderr.write(f"  [{path}: {type(e).__name__}]\n")


def _iter_xlsx(path: Path):
    try:
        import openpyxl
    except ImportError:
        sys.stderr.write(f"  [openpyxl not installed; skip {path}]\n")
        return
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as e:
        sys.stderr.write(f"  [{path}: {type(e).__name__}: {e}]\n")
        return
    global_idx = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Skip chartsheets — they have no row data, just chart objects.
        if not hasattr(ws, "iter_rows"):
            continue
        rows = ws.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            continue
        headers = [str(h) if h is not None else "" for h in headers]
        for row in rows:
            text = _row_to_text(headers, row)
            if text.strip():
                yield global_idx, f"[{sheet_name}] {text}"
                global_idx += 1


def _iter_parquet(path: Path):
    try:
        import pyarrow.parquet as pq
    except ImportError:
        sys.stderr.write(f"  [pyarrow not installed; skip {path}]\n")
        return
    try:
        table = pq.read_table(str(path))
    except Exception as e:
        sys.stderr.write(f"  [{path}: {type(e).__name__}: {e}]\n")
        return
    headers = table.column_names
    for i, row in enumerate(table.to_pylist()):
        text = _row_to_text(headers, [row.get(h) for h in headers])
        if text.strip():
            yield i, text


def _iter_ods(path: Path):
    try:
        from odf.opendocument import load
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
    except ImportError:
        sys.stderr.write(f"  [odfpy not installed; skip {path}]\n")
        return
    try:
        doc = load(str(path))
    except Exception as e:
        sys.stderr.write(f"  [{path}: {type(e).__name__}: {e}]\n")
        return

    def cell_text(cell):
        parts = []
        for p in cell.getElementsByType(P):
            parts.append(str(p))
        return " ".join(parts).strip()

    global_idx = 0
    for table in doc.getElementsByType(Table):
        rows = [tr for tr in table.getElementsByType(TableRow)]
        if not rows:
            continue
        headers = [cell_text(c) for c in rows[0].getElementsByType(TableCell)]
        for row in rows[1:]:
            cells = [cell_text(c) for c in row.getElementsByType(TableCell)]
            text = _row_to_text(headers, cells)
            if text.strip():
                yield global_idx, text
                global_idx += 1


class _TabularPlugin:
    extensions = (".csv", ".tsv", ".xlsx", ".ods", ".parquet", ".pq")
    source_format = "tabular"
    requires = ("openpyxl", "pyarrow", "odfpy")
    system_deps = ()
    default_on = True
    disable_env = "FORGE_DISABLE_TABULAR"

    def iter_chunks(self, path: Path, section: str, base_id: str, options: dict) -> Iterator[dict]:
        ext = path.suffix.lower()
        if ext == ".csv":
            src_fmt = "csv"
            iterator = _iter_csv(path, ",", src_fmt)
        elif ext == ".tsv":
            src_fmt = "tsv"
            iterator = _iter_csv(path, "\t", src_fmt)
        elif ext == ".xlsx":
            src_fmt = "xlsx"
            iterator = _iter_xlsx(path)
        elif ext == ".ods":
            src_fmt = "ods"
            iterator = _iter_ods(path)
        elif ext in (".parquet", ".pq"):
            src_fmt = "parquet"
            iterator = _iter_parquet(path)
        else:
            return

        for i, text in iterator:
            text = clean_text(text)
            if not text:
                continue
            yield {
                "id": f"{base_id}-r{i:06d}",
                "text": text,
                "format": "pretrain",
                "metadata": {
                    "source_file": str(path),
                    "source_format": src_fmt,
                    "section": section,
                    "doc_title": path.stem,
                    "chunk_type": "row",
                    "chunk_idx": i,
                    "char_count": len(text),
                },
            }


PLUGIN = _TabularPlugin()
